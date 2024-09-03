import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# Загрузка данных из JSON-файла
with open('/Users/zeinaddinzurgambayev/Desktop/parser/nurbek/count_brand_entries_five.json', 'r', encoding='utf-8') as file:
    data = json.load(file)

# Фильтрация данных: включаем только записи, где total_entries > 500
filtered_data = {brand: info for brand, info in data.items() if info['total_entries'] > 500}

# Создание нового Excel файла
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Filtered Brands"

# Установка стилей
header_font = Font(bold=True, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center")
header_fill = openpyxl.styles.PatternFill("solid", fgColor="4F81BD")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Запись заголовков с форматированием
headers = ['№', 'Бренд', 'Количество записей', 'Историческое количество товаров']
ws.append(headers)

for col_num, column_title in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = header_font
    cell.alignment = header_alignment
    cell.fill = header_fill
    cell.border = border

# Запись данных с нумерацией строк
for idx, (brand, info) in enumerate(filtered_data.items(), start=1):
    ws.append([idx, brand, info['total_entries'], info['item_count_historical']])

# Применение границ ко всем ячейкам
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Настройка ширины столбцов
column_widths = [5, 30, 20, 30]
for i, column_width in enumerate(column_widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width

# Сохранение Excel файла на рабочий стол
wb.save('/Users/zeinaddinzurgambayev/Desktop/parser/nurbek/xlsx_table/filtered_brand_entries5.xlsx')

print("Отфильтрованные данные записаны в filtered_brand_entries5.xlsx")
